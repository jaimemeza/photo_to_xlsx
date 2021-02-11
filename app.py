from PIL import Image
import openpyxl
import requests
from io import BytesIO
import math

wb = openpyxl.Workbook()
sheet = wb.active
filename = 'test'
url = 'https://i.ytimg.com/vi/AwI0kboH0sI/maxresdefault.jpg'
response = requests.get(url)
im = Image.open(BytesIO(response.content))

# Convert image to RGB if necessary
if not im.mode == "RGB":
    im = im.convert('RGB')

max_pixels = 200000
width, height = im.size
total_pixels = width * height

# Resize image if necessary
if total_pixels > max_pixels:
    ratio = height / width 
    width = round(math.sqrt(max_pixels/ratio))
    height = round(ratio * width)
    im = im.resize((width,height))

# Extract pixel values
pixel_values = list(im.getdata())

im.show()
height_mod = 1
for i, colors in enumerate(pixel_values):
    width_mod = (i % width) + 1
    if width_mod == 1 and i >= width:
        height_mod += 3
    for j, color in enumerate(colors):
        # Redimension each columns
        if height_mod == 1:
            sheet.column_dimensions[openpyxl.utils.get_column_letter(width_mod)].width = 1.7
        # Redimension each rows
        if width_mod == 2:
            sheet.row_dimensions[height_mod+j].height = 3
        array = ['00','00','00']
        array[j] = format(color,'02X')
        hex_value = ''.join(array)
        sheet.cell(row=height_mod+j, column=width_mod).fill = openpyxl.styles.PatternFill(start_color=hex_value,fill_type='solid')
        
wb.save(filename + '.xlsx')